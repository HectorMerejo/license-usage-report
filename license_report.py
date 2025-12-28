import os
import requests
from msal import ConfidentialClientApplication
from dotenv import load_dotenv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Load environment variables
load_dotenv()

CLIENT_ID = os.getenv('CLIENT_ID')
CLIENT_SECRET = os.getenv('CLIENT_SECRET')
TENANT_ID = os.getenv('TENANT_ID')

AUTHORITY = f'https://login.microsoftonline.com/{TENANT_ID}'
SCOPE = ['https://graph.microsoft.com/.default']

def get_access_token():
    """Authenticate and get access token"""
    print("Authenticating...")
    
    app = ConfidentialClientApplication(
        CLIENT_ID,
        authority=AUTHORITY,
        client_credential=CLIENT_SECRET
    )
    
    result = app.acquire_token_for_client(scopes=SCOPE)
    
    if 'access_token' in result:
        print("✓ Authentication successful\n")
        return result['access_token']
    else:
        print(f"✗ Authentication failed: {result.get('error_description')}\n")
        return None

def get_subscribed_skus(token):
    """Get all license SKUs (subscriptions)"""
    print("Retrieving license subscriptions...")
    
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }
    
    url = 'https://graph.microsoft.com/v1.0/subscribedSkus'
    
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        skus = response.json()['value']
        print(f"✓ Retrieved {len(skus)} license types\n")
        return skus
    except requests.exceptions.RequestException as e:
        print(f"✗ Error retrieving licenses: {e}\n")
        return []

def get_users_with_licenses(token):
    """Get all users and their assigned licenses"""
    print("Retrieving users and their licenses...")
    
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }
    
    # Select only fields we need to reduce data transfer
    url = 'https://graph.microsoft.com/v1.0/users?$select=displayName,userPrincipalName,assignedLicenses'
    
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        users = response.json()['value']
        print(f"✓ Retrieved {len(users)} users\n")
        return users
    except requests.exceptions.RequestException as e:
        print(f"✗ Error retrieving users: {e}\n")
        return []

def get_friendly_license_name(sku_part_number):
    """Convert SKU codes to friendly names"""
    # Common Microsoft license names
    license_names = {
        'ENTERPRISEPACK': 'Office 365 E3',
        'ENTERPRISEPREMIUM': 'Office 365 E5',
        'SPE_E3': 'Microsoft 365 E3',
        'SPE_E5': 'Microsoft 365 E5',
        'STANDARDPACK': 'Office 365 E1',
        'DESKLESSPACK': 'Office 365 F3',
        'EXCHANGESTANDARD': 'Exchange Online (Plan 1)',
        'EXCHANGEENTERPRISE': 'Exchange Online (Plan 2)',
        'SHAREPOINTSTANDARD': 'SharePoint Online (Plan 1)',
        'SHAREPOINTENTERPRISE': 'SharePoint Online (Plan 2)',
        'POWER_BI_STANDARD': 'Power BI (Free)',
        'POWER_BI_PRO': 'Power BI Pro',
        'PROJECTPROFESSIONAL': 'Project Plan 3',
        'VISIOCLIENT': 'Visio Plan 2',
        'TEAMS_EXPLORATORY': 'Microsoft Teams Exploratory',
        'FLOW_FREE': 'Power Automate Free',
        'POWERAPPS_VIRAL': 'Power Apps Trial',
    }
    
    return license_names.get(sku_part_number, sku_part_number)

def analyze_licenses(skus, users):
    """Analyze license usage"""
    print("Analyzing license usage...")
    
    license_data = []
    
    for sku in skus:
        sku_id = sku['skuId']
        sku_part_number = sku['skuPartNumber']
        friendly_name = get_friendly_license_name(sku_part_number)
        
        # Get totals
        enabled = sku['prepaidUnits']['enabled']
        consumed = sku['consumedUnits']
        available = enabled - consumed
        
        # Find users with this license
        users_with_license = []
        for user in users:
            if any(license['skuId'] == sku_id for license in user.get('assignedLicenses', [])):
                users_with_license.append(user['displayName'])
        
        license_data.append({
            'license_name': friendly_name,
            'sku_code': sku_part_number,
            'total_licenses': enabled,
            'assigned': consumed,
            'available': available,
            'utilization_pct': (consumed / enabled * 100) if enabled > 0 else 0,
            'users': users_with_license
        })
    
    print(f"✓ Analysis complete\n")
    return license_data

def create_excel_report(license_data):
    """Create formatted Excel report"""
    print("Creating Excel report...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "License Summary"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = ['License Type', 'SKU Code', 'Total Purchased', 'Assigned', 'Available', 'Utilization %']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for lic in license_data:
        ws.append([
            lic['license_name'],
            lic['sku_code'],
            lic['total_licenses'],
            lic['assigned'],
            lic['available'],
            f"{lic['utilization_pct']:.1f}%"
        ])
    
    # Color code utilization
    for row in range(2, len(license_data) + 2):
        util_cell = ws.cell(row=row, column=6)
        util_value = license_data[row-2]['utilization_pct']
        
        if util_value < 50:
            util_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
        elif util_value < 80:
            util_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        else:
            util_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Create detailed sheet with user assignments
    ws2 = wb.create_sheet("User Assignments")
    ws2.append(['License Type', 'User Name'])
    ws2[1][0].fill = header_fill
    ws2[1][0].font = header_font
    ws2[1][1].fill = header_fill
    ws2[1][1].font = header_font
    
    for lic in license_data:
        for user in lic['users']:
            ws2.append([lic['license_name'], user])
    
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 40
    
    # Save file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'license_report_{timestamp}.xlsx'
    wb.save(filename)
    
    print(f"✓ Report saved to: {filename}\n")
    return filename

def print_summary(license_data):
    """Print summary to console"""
    print("=" * 70)
    print("LICENSE USAGE SUMMARY")
    print("=" * 70)
    
    total_licenses = sum(lic['total_licenses'] for lic in license_data)
    total_assigned = sum(lic['assigned'] for lic in license_data)
    total_available = sum(lic['available'] for lic in license_data)
    
    print(f"Total Licenses Purchased:  {total_licenses}")
    print(f"Total Assigned:            {total_assigned}")
    print(f"Total Available (Unused):  {total_available}")
    print(f"Overall Utilization:       {(total_assigned/total_licenses*100):.1f}%")
    print()
    
    # Show licenses with low utilization (potential waste)
    print("⚠️  LOW UTILIZATION LICENSES (Potential Cost Savings):")
    print("-" * 70)
    
    low_util = [lic for lic in license_data if lic['utilization_pct'] < 70 and lic['total_licenses'] > 0]
    
    if low_util:
        for lic in sorted(low_util, key=lambda x: x['available'], reverse=True):
            print(f"{lic['license_name']:40} | {lic['available']:3} unused ({lic['utilization_pct']:.0f}% utilization)")
    else:
        print("No licenses with low utilization found.")
    
    print("=" * 70)

def main():
    print("\n" + "=" * 70)
    print("MICROSOFT LICENSE USAGE REPORT GENERATOR")
    print("=" * 70 + "\n")
    
    # Step 1: Authenticate
    token = get_access_token()
    if not token:
        print("Failed to authenticate. Check your .env file credentials.")
        return
    
    # Step 2: Get license subscriptions (SKUs)
    skus = get_subscribed_skus(token)
    if not skus:
        print("No licenses found or error retrieving licenses.")
        return
    
    # Step 3: Get users and their license assignments
    users = get_users_with_licenses(token)
    if not users:
        print("No users found or error retrieving users.")
        return
    
    # Step 4: Analyze license usage
    license_data = analyze_licenses(skus, users)
    
    # Step 5: Create Excel report
    create_excel_report(license_data)
    
    # Step 6: Print summary
    print_summary(license_data)

if __name__ == "__main__":
    main()